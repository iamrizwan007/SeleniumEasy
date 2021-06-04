from selenium import webdriver
import logging
import inspect
import openpyxl


class BaseClass:
    def __init__(self):
        pass

    @staticmethod
    def getCustomLogger(level):
        function_name = inspect.stack()[1][3]
        logger_name = function_name + "logger"

        logger = logging.getLogger(logger_name)
        logger.setLevel(level)

        handler = logging.FileHandler(filename="C:\\Users\\Rizwan\\PycharmProjects\\SeleniumEasy\\Logs\\testlogs.log", mode='w')
        handler.setLevel(level)

        formatter = logging.Formatter('%(asctime)s %(levelname)s %(name)s %(message)s', datefmt='%d/%m/%Y %I:%M:%S %p')
        handler.setFormatter(formatter)

        logger.addHandler(handler)
        return logger

    @staticmethod
    def get_data_from_excel(path):
        workbook = openpyxl.load_workbook(filename=path)
        sheet = workbook.active
        row = sheet.max_row
        col = sheet.max_column
        header = []
        for r in range(1, 2):
            for c in range(1, col + 1):
                header.append(sheet.cell(r, c).value)
        dict_list = []
        for r in range(2, row + 1):
            for c in range(1, col + 1):
                my_dict = {}
                my_dict[header[c - 1]] = sheet.cell(r, c).value
                dict_list.append(my_dict)
        return dict_list
