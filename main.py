import openpyxl


workbook = openpyxl.load_workbook(filename="C:\\Users\\Rizwan\\Desktop\\Floridabar_Extract.xlsx")
sheet = workbook.active
row = sheet.max_row
col = sheet.max_column
print(row,col)
header = []
for r in range(1,2):
    for c in range(1,col+1):
        print(sheet.cell(r,c).value)
        header.append(sheet.cell(r,c).value)
print(header)
td = []
for r in range(2, row+1):
    d = {}
    for c in range(1, col+1):
        my_dict = {}
        my_dict[header[c-1]] = sheet.cell(r,c).value
        td.append(my_dict)
print(td)

    def allurereport():
import logging

import allure
from allure_commons.types import AttachmentType
from selenium import webdriver
from page_objects.main_page import MainPageObjects
from page_objects.simple_form_objects import SimpleFormPageObjects
from Utilities.Base import BaseClass
import pytest

@pytest.mark.usefixtures("setup")
class TestInputForms:
    @allure.severity(allure.severity_level.CRITICAL)
    def test_simple_form_demo(self):
        mylogger = BaseClass.getCustomLogger(logging.INFO)
        mylogger.info("logger created")
        main_page_obj = MainPageObjects(self.driver)
        simple_form_page_obj = SimpleFormPageObjects(self.driver)
        main_page_obj.expand_input_forms()
        mylogger.info("Expanded input forms")
        main_page_obj.click_simple_form_demo()
        mylogger.info("clicked on simple form demo")
        main_page_obj.verify_title()
        mylogger.info("Title verified")
        msg = "testing message"
        simple_form_page_obj.single_input_message(msg)
        simple_form_page_obj.verify_single_input_message(msg)
        num1 = 10
        num2 = 20
        simple_form_page_obj.single_input_num1(num1)
        simple_form_page_obj.single_input_num2(num2)
        simple_form_page_obj.click_on_total()
        simple_form_page_obj.verify_total_sum(num1, num2)
    @allure.severity(allure.severity_level.CRITICAL)
    @pytest.mark.skip
    def test_checkbox_demo(self):
        pass

    @allure.severity(allure.severity_level.BLOCKER)
    def test_fail(self):
        allure.attach(self.driver.get_screenshot_as_png(),name='testfailscreen',attachment_type= AttachmentType.PNG)
        assert False
